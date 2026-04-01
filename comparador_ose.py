"""
Comparador OSE - SANEPAR / 2S Engenharia e Geotecnologia
=========================================================
Compara informações entre MAPA, PERFIL e PLANILHA (MND) de OSEs de rede coletora.

Estrutura de pastas:
    pasta_raiz/
        MAPAS/       OSE-552.pdf, OSE-553.pdf ...
        PERFIS/      OSE-552.pdf, OSE-553.pdf ...
        PLANILHAS/   OSE-552.pdf, OSE-553.pdf ...

Uso:
    python comparador_ose.py [pasta_raiz] [saida.xlsx]

    pasta_raiz: diretório que contém MAPAS, PERFIS, PLANILHAS  (padrão: diretório atual)
    saida.xlsx: relatório gerado  (padrão: pasta_raiz/relatorio_divergencias.xlsx)
"""

import os, re, sys
from collections import Counter
from dataclasses import dataclass, field
from typing import Optional
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOLERANCIA = 0.005   # 5 mm

# ---------------------------------------------------------------------------
# Estruturas de dados
# ---------------------------------------------------------------------------

@dataclass
class DadosOSE:
    fonte: str = ""
    numero_ose: str = ""
    pv_inicial: str = ""
    pv_final: str = ""
    extensao: Optional[float] = None
    declividade: Optional[float] = None
    diametro: Optional[int] = None
    prof_vala_ini: Optional[float] = None
    prof_vala_fin: Optional[float] = None
    ct_ini: Optional[float] = None
    cf_ini: Optional[float] = None
    ct_fin: Optional[float] = None
    cf_fin: Optional[float] = None
    erros_extracao: list = field(default_factory=list)

@dataclass
class Divergencia:
    campo: str
    fonte_a: str; valor_a: str
    fonte_b: str; valor_b: str
    diferenca: str = ""

# ---------------------------------------------------------------------------
# Utilitários
# ---------------------------------------------------------------------------

def _txt(caminho: str) -> str:
    texto = ""
    try:
        with pdfplumber.open(caminho) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    texto += t + "\n"
    except Exception as e:
        print(f"  [ERRO] {caminho}: {e}")
    return texto

def _f(s) -> Optional[float]:
    if s is None: return None
    try: return float(str(s).replace(',', '.').strip())
    except: return None

def _npv(s: str) -> str:
    return re.sub(r'((?:PV|TL))\s*[\-\s]?\s*(\d+)', r'\1-\2', s.strip().upper())

def _listar_oses(pasta: str) -> set:
    nums = set()
    if not os.path.isdir(pasta): return nums
    for f in os.listdir(pasta):
        m = re.search(r'OSE[-\s]?(\d+)', f, re.IGNORECASE)
        if m: nums.add(m.group(1).zfill(3))
    return nums

def _pdf_da_ose(pasta: str, numero: str) -> Optional[str]:
    if not os.path.isdir(pasta): return None
    num_int = str(int(numero))
    for f in os.listdir(pasta):
        if re.search(rf'OSE[-\s]?0*{num_int}\.pdf', f, re.IGNORECASE):
            return os.path.join(pasta, f)
    return None

# ---------------------------------------------------------------------------
# EXTRATOR — PLANILHA (MND)
# ---------------------------------------------------------------------------
# Texto pdfplumber:
#   "PARCERIA PÚBLICO PRIVADA OSE-552\n"
#   "PV-558 0,00 0,00 475,143 472,901 200 473,001 2,142 475,901 0,1219 144 3,6 2,242 0,758 PV...\n"
#   "TL-501 3,00 77,35 483,430 482,330 200 482,430 1,000 486,330 0,1219 144 3,6 1,100 2,900 TL...\n"
#   "DN 150mm - PVC\n"

def extrair_planilha(caminho: str) -> DadosOSE:
    dados = DadosOSE(fonte="PLANILHA")
    texto = _txt(caminho)
    if not texto:
        dados.erros_extracao.append("Não foi possível ler o PDF."); return dados

    m = re.search(r'OSE[\s\-]*(\d+)', texto, re.IGNORECASE)
    if m: dados.numero_ose = m.group(1).zfill(3)

    m = re.search(r'\bDN\s*(\d+)\s*mm', texto, re.IGNORECASE)
    if m: dados.diametro = int(m.group(1))

    pvs = []
    for linha in texto.splitlines():
        m = re.match(
            r'^((?:PV|TL)[\-\s]?\d+)\s+'
            r'([\d,\.]+)\s+([\d,\.]+)\s+'    # dist, dist_acum
            r'([\d,\.]+)\s+([\d,\.]+)\s+'    # c_terreno, c_fundo
            r'(\d+)\s+([\d,\.]+)\s+([\d,\.]+)\s+([\d,\.]+)\s+'  # diam, c_eixo, prof_eixo, c_regua
            r'([\d,\.]+)\s+(\d+)\s+([\d,\.]+)\s+([\d,\.]+)',    # decliv, diam_int, prof_vala, alt_regua
            linha.strip())
        if m:
            pvs.append({
                'nome':        _npv(m.group(1)),
                'dist_acum':   _f(m.group(3)),
                'c_terreno':   _f(m.group(4)),
                'c_fundo':     _f(m.group(5)),
                'declividade': _f(m.group(10)),
                'prof_vala':   _f(m.group(12)),
            })

    if not pvs:
        dados.erros_extracao.append("Nenhuma linha PV/TL encontrada na tabela."); return dados

    ini, fin = pvs[0], pvs[-1]
    dados.pv_inicial    = ini['nome'];     dados.pv_final       = fin['nome']
    dados.ct_ini        = ini['c_terreno']; dados.cf_ini         = ini['c_fundo']
    dados.ct_fin        = fin['c_terreno']; dados.cf_fin         = fin['c_fundo']
    dados.extensao      = fin['dist_acum']
    dados.prof_vala_ini = ini['prof_vala']; dados.prof_vala_fin  = fin['prof_vala']
    dados.declividade   = (pvs[1]['declividade'] if len(pvs) > 1 and pvs[1]['declividade']
                           else ini['declividade'])
    return dados

# ---------------------------------------------------------------------------
# EXTRATOR — MAPA
# ---------------------------------------------------------------------------
# Texto pdfplumber:
#   "Nº OSE: 552ESCALA: ..."                      → número (sem quebra de linha)
#   "TL-501 h:1.100\nCT:483.430\nCF:482.330\n"   → PV com cotas
#   "OSE – 552\nØ150mm\ni=0,1219\nL=77,35m"      → bloco da rede (cada dado em linha separada)
#
# ATENÇÃO: o mapa tem PVs de outras OSEs. Os da OSE-NNN ficam
# imediatamente antes e depois do bloco "OSE – NNN\nØ...\ni=...\nL=...".

def extrair_mapa(caminho: str, numero_ose: str = "") -> DadosOSE:
    dados = DadosOSE(fonte="MAPA")
    texto = _txt(caminho)
    if not texto:
        dados.erros_extracao.append("Não foi possível ler o PDF."); return dados

    # Número OSE — "Nº OSE: 552ESCALA" (compacto, sem \n entre número e próximo campo)
    m = re.search(r'N[ºo°]\s*OSE\s*[:\s]+(\d+)', texto, re.IGNORECASE)
    if m: dados.numero_ose = m.group(1).zfill(3)

    ose_num = numero_ose or dados.numero_ose
    num_int = str(int(ose_num)) if ose_num else ""
    pos_bloco = -1

    if num_int:
        # Padrão principal: cada dado em linha própria
        pat = (r'OSE\s*[–\-]\s*0*' + num_int + r'\s*\n'
               r'\s*[ØOo]?\s*(\d+)\s*mm\s*\n'    # grupo 1 = diâmetro
               r'\s*i\s*=\s*([\d,\.]+)\s*\n'      # grupo 2 = declividade
               r'\s*L\s*=\s*([\d,\.]+)\s*m')       # grupo 3 = extensão
        m = re.search(pat, texto, re.IGNORECASE)
        if m:
            dados.diametro    = int(m.group(1))
            dados.declividade = _f(m.group(2))
            dados.extensao    = _f(m.group(3))
            pos_bloco         = m.start()
        else:
            # Fallback: tudo na mesma linha
            pat2 = (r'[ØOo](\d+)mm[^\n]*OSE\s*[–\-]\s*0*' + num_int +
                    r'[^\n]*i\s*=\s*([\d,\.]+)[^\n]*L\s*=\s*([\d,\.]+)\s*m')
            m2 = re.search(pat2, texto, re.IGNORECASE)
            if m2:
                dados.diametro    = int(m2.group(1))
                dados.declividade = _f(m2.group(2))
                dados.extensao    = _f(m2.group(3))
                pos_bloco         = m2.start()
            else:
                dados.erros_extracao.append(f"Bloco OSE-{num_int} não encontrado no mapa.")

    # Coleta todos os PVs/TLs com CT e CF
    # Padrão: "PV-558 h:2.242\nCT:475.143\nCF:472.901"
    blocos_pv = re.findall(
        r'((?:PV|TL)[\-]?\d+)\s+h\s*[:\.]?\s*([\d,\.]+)\s*\n'
        r'CT\s*[:\.]?\s*([\d,\.]+)\s*\n'
        r'CF\s*[:\.]?\s*([\d,\.]+)',
        texto, re.IGNORECASE)

    if not blocos_pv:
        dados.erros_extracao.append("Nenhum PV/TL com CT/CF encontrado no mapa.")
        return dados

    pv_posicoes = []
    for b in blocos_pv:
        nome = _npv(b[0])
        ct, cf = _f(b[2]), _f(b[3])
        pos = texto.find(b[0])
        pv_posicoes.append((pos, nome, ct, cf))

    # Identifica os 2 PVs da OSE: imediatamente antes e depois do bloco
    if pos_bloco >= 0:
        antes  = [(p, n, ct, cf) for p, n, ct, cf in pv_posicoes if p < pos_bloco]
        depois = [(p, n, ct, cf) for p, n, ct, cf in pv_posicoes if p > pos_bloco]
        candidatos = []
        if antes:  candidatos.append(antes[-1])   # último antes do bloco
        if depois: candidatos.append(depois[0])   # primeiro depois do bloco
    else:
        # Sem posição: usa CF extrema
        ordenados = sorted(pv_posicoes, key=lambda x: x[3] or 0)
        candidatos = [ordenados[0], ordenados[-1]]

    if len(candidatos) >= 2:
        # Menor CF = montante (PV inicial), maior CF = jusante (PV final)
        candidatos.sort(key=lambda x: x[3] or 0)
        dados.pv_inicial = candidatos[0][1];  dados.ct_ini = candidatos[0][2]; dados.cf_ini = candidatos[0][3]
        dados.pv_final   = candidatos[-1][1]; dados.ct_fin = candidatos[-1][2]; dados.cf_fin = candidatos[-1][3]
    elif candidatos:
        _, nome, ct, cf = candidatos[0]
        dados.pv_inicial = nome; dados.ct_ini = ct; dados.cf_ini = cf
        dados.erros_extracao.append("Apenas 1 PV identificado no mapa para esta OSE.")

    return dados

# ---------------------------------------------------------------------------
# EXTRATOR — PERFIL
# ---------------------------------------------------------------------------
# Texto pdfplumber:
#   "Nº OSE: 552ESCALA: ..."
#   "COTA TOPO 485,430 ... 475,143"             (primeiro=CT_fin jusante, último=CT_ini montante)
#   "EXTENSÃO ACUMULADA 0,00m ... 77,35m"       (maior valor = extensão total)
#   "DECLIVIDADE (mm) 0,1219 0,1219 ..."        (moda = declividade do trecho)
#   "DIAMETRO TUBO Ø150mm Ø150mm ..."
#   "GI DO TUBO 482.330 ... 473.187"            (primeiro=CF_fin jusante, último=CF_ini montante)
#   "PV  TL-501 PIT-7290 ... PV-558"            (linha PV da tabela)
#
# CONVENÇÃO: tabela vai da ESQUERDA (jusante=PV final) para DIREITA (montante=PV inicial)

def extrair_perfil(caminho: str) -> DadosOSE:
    dados = DadosOSE(fonte="PERFIL")
    texto = _txt(caminho)
    if not texto:
        dados.erros_extracao.append("Não foi possível ler o PDF."); return dados

    # Número OSE — "Nº OSE: 552ESCALA" (compacto)
    m = re.search(r'N[ºo°]\s*OSE\s*[:\s]+(\d+)', texto, re.IGNORECASE)
    if m: dados.numero_ose = m.group(1).zfill(3)

    # Diâmetro — "DIAMETRO TUBO Ø150mm ..."
    m = re.search(r'DIAMETRO\s+TUBO\s+[ØOo]?(\d+)mm', texto, re.IGNORECASE)
    if m: dados.diametro = int(m.group(1))

    # Extensão — "EXTENSÃO ACUMULADA 0,00m ... 77,35m" — pega o MAIOR valor da linha
    m = re.search(r'EXTENS[ÃA]O\s+ACUMULADA\s+(.+)', texto, re.IGNORECASE)
    if m:
        todos = re.findall(r'([\d,\.]+)\s*m', m.group(1))
        floats = [_f(v) for v in todos if _f(v) is not None]
        if floats: dados.extensao = max(floats)

    # Declividade — linha "DECLIVIDADE (mm) 0,1219 ..." — usa MODA
    m = re.search(r'DECLIVIDADE\s*\(mm\)\s+(.+)', texto, re.IGNORECASE)
    if m:
        vals = re.findall(r'[\d,\.]+', m.group(1))
        floats = [_f(v) for v in vals if _f(v) and 0.001 < (_f(v) or 0) < 1.0]
        if floats: dados.declividade = Counter(floats).most_common(1)[0][0]

    # PVs — linha "PV  TL-501 PIT-7290 ... PV-558"
    # Esquerda = jusante = PV final; Direita = montante = PV inicial
    m_pv = re.search(r'^PV\s+((?:(?:PV|TL|PIT)[\-\s]?\d+\s*)+)', texto, re.MULTILINE | re.IGNORECASE)
    if m_pv:
        # Filtra apenas PV/TL (ignora PIT intermediários)
        pvs_tab = re.findall(r'((?:PV|TL)[\-]?\d+)', m_pv.group(1), re.IGNORECASE)
        if pvs_tab:
            dados.pv_final   = _npv(pvs_tab[0])   # esquerda = jusante = final
            dados.pv_inicial = _npv(pvs_tab[-1])  # direita = montante = inicial
    else:
        # Fallback: qualquer PV/TL no texto (excluindo PITs)
        todos = re.findall(r'(?<!\w)((?:PV|TL)[\-]?\d+)(?!\w)', texto, re.IGNORECASE)
        uniq = list(dict.fromkeys([_npv(p) for p in todos]))
        if uniq:
            dados.pv_final   = uniq[0]
            dados.pv_inicial = uniq[-1]

    # Cotas de topo — "COTA TOPO 485,430 ... 475,143"
    # primeiro = CT_fin (jusante), último = CT_ini (montante)
    m = re.search(r'COTA\s+TOPO\s+(.+)', texto, re.IGNORECASE)
    if m:
        cotas = re.findall(r'\d{3}[,\.]\d{3}', m.group(1))
        if cotas:
            dados.ct_fin = _f(cotas[0])
            dados.ct_ini = _f(cotas[-1])

    # Cotas de fundo — "GI DO TUBO 482.330 ... 473.187"
    # primeiro = CF_fin (jusante), último = CF_ini (montante)
    m = re.search(r'GI\s+DO\s+TUBO\s+(.+)', texto, re.IGNORECASE)
    if m:
        cotas = re.findall(r'\d{3}\.\d{3}', m.group(1))
        if cotas:
            dados.cf_fin = _f(cotas[0])
            dados.cf_ini = _f(cotas[-1])

    if not dados.extensao:    dados.erros_extracao.append("Extensão não encontrada.")
    if not dados.declividade: dados.erros_extracao.append("Declividade não encontrada.")
    if not dados.pv_inicial:  dados.erros_extracao.append("PV inicial não encontrado.")

    return dados

# ---------------------------------------------------------------------------
# COMPARAÇÃO
# ---------------------------------------------------------------------------

def comparar(planilha: DadosOSE, mapa: DadosOSE, perfil: DadosOSE) -> list:
    divs = []
    fontes = [("PLANILHA", planilha), ("MAPA", mapa), ("PERFIL", perfil)]

    def chk(campo, getter):
        valores = {n: getter(d) for n, d in fontes
                   if getter(d) is not None and getter(d) != "" and getter(d) != 0}
        nomes = list(valores.keys())
        for i in range(len(nomes)):
            for j in range(i + 1, len(nomes)):
                na, nb = nomes[i], nomes[j]
                va, vb = valores[na], valores[nb]
                if isinstance(va, float) and isinstance(vb, float):
                    diff = abs(va - vb)
                    if diff > TOLERANCIA:
                        divs.append(Divergencia(campo=campo,
                            fonte_a=na, valor_a=f"{va:.3f}",
                            fonte_b=nb, valor_b=f"{vb:.3f}",
                            diferenca=f"{diff:.3f} m"))
                elif isinstance(va, int) and isinstance(vb, int):
                    if va != vb:
                        divs.append(Divergencia(campo=campo,
                            fonte_a=na, valor_a=str(va), fonte_b=nb, valor_b=str(vb)))
                elif str(va).strip().upper() != str(vb).strip().upper():
                    divs.append(Divergencia(campo=campo,
                        fonte_a=na, valor_a=str(va), fonte_b=nb, valor_b=str(vb)))

    chk("PV Inicial",            lambda d: d.pv_inicial or None)
    chk("PV Final",              lambda d: d.pv_final or None)
    chk("Extensão total (m)",    lambda d: d.extensao)
    chk("Declividade",           lambda d: d.declividade)
    chk("Diâmetro DN (mm)",      lambda d: float(d.diametro) if d.diametro else None)
    chk("CT PV Inicial (m)",     lambda d: d.ct_ini)
    chk("CF PV Inicial (m)",     lambda d: d.cf_ini)
    chk("CT PV Final (m)",       lambda d: d.ct_fin)
    chk("CF PV Final (m)",       lambda d: d.cf_fin)
    chk("Prof. Vala PV Ini (m)", lambda d: d.prof_vala_ini)
    chk("Prof. Vala PV Fin (m)", lambda d: d.prof_vala_fin)

    return divs

# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------

_CH = "1F3864"; _CT = "2E75B6"; _OK = "E2EFDA"; _DV = "FFE0E0"
_WN = "FFF2CC"; _P  = "F5F5F5"; _B  = "FFFFFF"

def _bd():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _c(ws, r, col, v="", bold=False, bg=None, fg="000000", cx=False, sz=10):
    c = ws.cell(row=r, column=col, value=v)
    c.font = Font(name='Arial', bold=bold, color=fg, size=sz)
    c.border = _bd()
    if bg: c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal='center' if cx else 'left', vertical='center')
    return c

def _sec(ws, r, nc, txt, bg=_CT):
    ws.merge_cells(f"A{r}:{get_column_letter(nc)}{r}")
    cc = ws[f"A{r}"]
    cc.value = txt
    cc.font = Font(name='Arial', bold=True, size=11, color="FFFFFF")
    cc.fill = PatternFill("solid", start_color=bg)
    cc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 20
    return r + 1


def gerar_aba_ose(wb, num, planilha, mapa, perfil, divs):
    ws = wb.create_sheet(title=f"OSE-{num}")
    NC = 8

    ws.merge_cells(f"A1:{get_column_letter(NC)}1")
    c = ws["A1"]; c.value = f"COMPARATIVO OSE-{num}"
    c.font = Font(name='Arial', bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=_CH)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    nd = len(divs)
    st = "✅  SEM DIVERGÊNCIAS" if nd == 0 else f"⚠️   {nd} DIVERGÊNCIA(S) ENCONTRADA(S)"
    cf = _OK if nd == 0 else _DV; ff = "1A5C2A" if nd == 0 else "8B0000"
    ws.merge_cells(f"A2:{get_column_letter(NC)}2")
    c = ws["A2"]; c.value = st
    c.font = Font(name='Arial', bold=True, size=11, color=ff)
    c.fill = PatternFill("solid", start_color=cf)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    row = 4
    row = _sec(ws, row, NC, "DADOS EXTRAÍDOS POR FONTE")
    for col, cab in zip([1,2,4,6], ["Campo","PLANILHA (MND)","MAPA","PERFIL"]):
        _c(ws, row, col, cab, bold=True, bg="BDD7EE", cx=True)
    ws.row_dimensions[row].height = 18; row += 1

    campos = [
        ("Número OSE",         lambda d: d.numero_ose),
        ("PV Inicial",         lambda d: d.pv_inicial),
        ("PV Final",           lambda d: d.pv_final),
        ("Extensão (m)",       lambda d: f"{d.extensao:.3f}" if d.extensao else "-"),
        ("Declividade",        lambda d: f"{d.declividade:.4f}" if d.declividade else "-"),
        ("Diâmetro DN (mm)",   lambda d: str(d.diametro) if d.diametro else "-"),
        ("CT PV Inicial (m)",  lambda d: f"{d.ct_ini:.3f}" if d.ct_ini else "-"),
        ("CF PV Inicial (m)",  lambda d: f"{d.cf_ini:.3f}" if d.cf_ini else "-"),
        ("CT PV Final (m)",    lambda d: f"{d.ct_fin:.3f}" if d.ct_fin else "-"),
        ("CF PV Final (m)",    lambda d: f"{d.cf_fin:.3f}" if d.cf_fin else "-"),
        ("Prof. Vala Ini (m)", lambda d: f"{d.prof_vala_ini:.3f}" if d.prof_vala_ini else "-"),
        ("Prof. Vala Fin (m)", lambda d: f"{d.prof_vala_fin:.3f}" if d.prof_vala_fin else "-"),
    ]

    for idx, (label, getter) in enumerate(campos):
        bg = _P if idx % 2 == 0 else _B
        _c(ws, row, 1, label, bold=True, bg=bg)
        for col, fonte in zip([2, 4, 6], [planilha, mapa, perfil]):
            val = getter(fonte) if fonte else "-"
            _c(ws, row, col, val or "-", bg=bg, cx=True)
        ws.row_dimensions[row].height = 16; row += 1

    alertas = [f"[{f.fonte}] {e}" for f in [planilha, mapa, perfil] if f for e in f.erros_extracao]
    if alertas:
        row += 1
        row = _sec(ws, row, NC, "⚠️  ALERTAS DE EXTRAÇÃO", bg="C55A11")
        for al in alertas:
            ws.merge_cells(f"A{row}:{get_column_letter(NC)}{row}")
            c = ws[f"A{row}"]; c.value = al
            c.font = Font(name='Arial', size=9, color="8B0000")
            c.fill = PatternFill("solid", start_color="FDECEA")
            c.alignment = Alignment(vertical='center', wrap_text=True)
            ws.row_dimensions[row].height = 16; row += 1

    row += 1
    row = _sec(ws, row, NC, "DIVERGÊNCIAS DETECTADAS")
    if not divs:
        ws.merge_cells(f"A{row}:{get_column_letter(NC)}{row}")
        c = ws[f"A{row}"]; c.value = "Nenhuma divergência encontrada entre os três documentos."
        c.font = Font(name='Arial', size=10, color="1A5C2A")
        c.fill = PatternFill("solid", start_color=_OK)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 18
    else:
        for i, cab in enumerate(["Campo","Fonte A","Valor A","Fonte B","Valor B","Diferença"], 1):
            _c(ws, row, i, cab, bold=True, bg="F4B942", cx=True)
        ws.row_dimensions[row].height = 18; row += 1
        for idx, div in enumerate(divs):
            bg = _DV if idx % 2 == 0 else "FFD0D0"
            for i, v in enumerate([div.campo, div.fonte_a, div.valor_a,
                                    div.fonte_b, div.valor_b, div.diferenca], 1):
                _c(ws, row, i, v, bg=bg, cx=(i != 1))
            ws.row_dimensions[row].height = 16; row += 1

    for i, w in enumerate([28, 18, 4, 18, 4, 18, 4, 4], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def gerar_aba_resumo(wb, resultados):
    ws = wb.create_sheet(title="RESUMO", index=0)
    NC = 6
    ws.merge_cells(f"A1:{get_column_letter(NC)}1")
    c = ws["A1"]; c.value = "RELATÓRIO DE DIVERGÊNCIAS — OSE SANEPAR / 2S ENGENHARIA"
    c.font = Font(name='Arial', bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=_CH)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    row = 3
    for i, cab in enumerate(["OSE","PV Inicial","PV Final","Extensão (m)","Divergências","Status"], 1):
        _c(ws, row, i, cab, bold=True, bg=_CT, fg="FFFFFF", cx=True)
    ws.row_dimensions[row].height = 20; row += 1

    for num, plan, mapa, perf, divs in resultados:
        nd = len(divs)
        bg = _OK if nd == 0 else (_DV if nd > 2 else _WN)
        st = "✅ OK" if nd == 0 else f"⚠️  {nd} div."
        ref = plan or mapa or perf
        ext = ref.extensao if ref else None
        for i, v in enumerate([f"OSE-{num}", ref.pv_inicial if ref else "-",
                                ref.pv_final if ref else "-",
                                f"{ext:.2f}" if ext else "-", str(nd), st], 1):
            _c(ws, row, i, v, bg=bg, cx=(i != 2))
        ws.row_dimensions[row].height = 16; row += 1

    for i, w in enumerate([12, 16, 16, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

# ---------------------------------------------------------------------------
# AUDITORIA DE ARQUIVOS
# ---------------------------------------------------------------------------

def verificar_faltantes(pasta_mapas: str, pasta_perfis: str, pasta_planilhas: str) -> dict:
    """Retorna quais arquivos (pelo nome base) faltam em cada pasta."""
    def _nomes(pasta: str) -> set:
        if not pasta or not os.path.isdir(pasta):
            return set()
        return {os.path.splitext(f)[0] for f in os.listdir(pasta)
                if os.path.isfile(os.path.join(pasta, f))}

    nomes_map = _nomes(pasta_mapas)
    nomes_per = _nomes(pasta_perfis)
    nomes_pla = _nomes(pasta_planilhas)
    todos     = sorted(nomes_map | nomes_per | nomes_pla)

    return {
        "todos":          todos,
        "mapa":           sorted(nomes_map),
        "perfil":         sorted(nomes_per),
        "planilha":       sorted(nomes_pla),
        "falta_mapa":     sorted((nomes_per | nomes_pla) - nomes_map),
        "falta_perfil":   sorted((nomes_map | nomes_pla) - nomes_per),
        "falta_planilha": sorted((nomes_map | nomes_per) - nomes_pla),
    }


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def processar_pastas(pasta_mapas: str, pasta_perfis: str, pasta_planilhas: str,
                     arquivo_saida: str):
    p_map = pasta_mapas
    p_per = pasta_perfis
    p_pla = pasta_planilhas

    for p in [p_map, p_per, p_pla]:
        if not os.path.isdir(p):
            print(f"[AVISO] Pasta não encontrada: {p}")

    oses = set()
    for p in [p_map, p_per, p_pla]:
        oses.update(_listar_oses(p))

    if not oses:
        print("[ERRO] Nenhum arquivo OSE-NNN.pdf encontrado nas subpastas."); return

    print(f"\nOSEs encontrados: {', '.join(sorted(oses))}\n")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    resultados = []

    for num in sorted(oses):
        print(f"Processando OSE-{num}...")
        arq_pla = _pdf_da_ose(p_pla, num)
        arq_map = _pdf_da_ose(p_map, num)
        arq_per = _pdf_da_ose(p_per, num)

        planilha = extrair_planilha(arq_pla) if arq_pla else DadosOSE(fonte="PLANILHA", erros_extracao=["Arquivo não encontrado"])
        mapa     = extrair_mapa(arq_map, num) if arq_map else DadosOSE(fonte="MAPA",     erros_extracao=["Arquivo não encontrado"])
        perfil   = extrair_perfil(arq_per)    if arq_per else DadosOSE(fonte="PERFIL",   erros_extracao=["Arquivo não encontrado"])

        for d in [planilha, mapa, perfil]:
            if not d.numero_ose: d.numero_ose = num

        divs = comparar(planilha, mapa, perfil)
        print(f"  {'✅' if not divs else '⚠️ '} {len(divs)} divergência(s)")
        gerar_aba_ose(wb, num, planilha, mapa, perfil, divs)
        resultados.append((num, planilha, mapa, perfil, divs))

    gerar_aba_resumo(wb, resultados)
    wb.save(arquivo_saida)
    print(f"\n✅ Relatório salvo em: {arquivo_saida}")
    return resultados


def processar(pasta_raiz: str, arquivo_saida: str):
    """Mantém compatibilidade: root contendo MAPAS/, PERFIS/, PLANILHAS/."""
    return processar_pastas(
        os.path.join(pasta_raiz, "MAPAS"),
        os.path.join(pasta_raiz, "PERFIS"),
        os.path.join(pasta_raiz, "PLANILHAS"),
        arquivo_saida,
    )


if __name__ == "__main__":
    pasta = sys.argv[1] if len(sys.argv) > 1 else "."
    saida = sys.argv[2] if len(sys.argv) > 2 else os.path.join(pasta, "relatorio_divergencias.xlsx")
    processar(pasta, saida)
