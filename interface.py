"""
Interface — OSEAudit v1.0
A2Z Projetos em parceria com 2S Engenharia e Geotecnologia
"""

import contextlib
import io
import json
import os
import subprocess
import sys
import tempfile
import threading
import tkinter as tk
import urllib.error
import urllib.request
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

# ── Resource helper ────────────────────────────────────────────────────────
def _res(*parts):
    base = (os.path.dirname(sys.executable)
            if getattr(sys, "frozen", False)
            else os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, *parts)

# ── Core imports ───────────────────────────────────────────────────────────
try:
    from comparador_ose import (processar_pastas, gerar_aba_ose,
                                 gerar_aba_resumo, verificar_faltantes)
except ImportError as e:
    tk.Tk().withdraw()
    messagebox.showerror("Erro", f"Módulo comparador não encontrado:\n{e}")
    sys.exit(1)

from splash import SplashScreen

# ── Constants ──────────────────────────────────────────────────────────────
NOME_PROGRAMA = "OSEAudit"
VERSAO        = "1.0"
SUBTITULO     = "Comparação e Auditoria de Documentos OSE"
GITHUB_REPO   = "A2ZPROJ/OSEAudit"

# ── Color Palette — GitHub Dark inspired ──────────────────────────────────
BG       = "#0D1117"    # Main window background
SURFACE  = "#161B22"    # Elevated surface
CARD     = "#1C2128"    # Cards / inputs
OVERLAY  = "#21262D"    # Hover state

BORDER   = "#30363D"    # Default border
BORDER_A = "#6E7681"    # Active / focus border

RED      = "#DA3633"    # Primary brand red
RED_H    = "#B91C1C"    # Red hover
RED_D    = "#3D0A0A"    # Deep red background
RED_L    = "#FF7B72"    # Light red (text)

TXT      = "#E6EDF3"    # Primary text
TXT2     = "#8B949E"    # Secondary text
TXT3     = "#484F58"    # Muted/dim text

GRN      = "#3FB950"    # Success
GRN_BG   = "#0D2F1F"    # Success background
AMB      = "#D29922"    # Warning
AMB_BG   = "#2A1F00"    # Warning background
ERR      = "#F85149"    # Error
BLU      = "#58A6FF"    # Info

LOG_BG   = "#010409"    # Terminal background
LOG_FG   = "#C9D1D9"    # Terminal text

F_UI     = ("Segoe UI", 10)
F_BOLD   = ("Segoe UI", 10, "bold")
F_SMALL  = ("Segoe UI",  8)
F_MED    = ("Segoe UI",  9)
F_MONO   = ("Consolas",  9)
F_CAP    = ("Segoe UI",  8, "bold")


# ── Flat button factory ────────────────────────────────────────────────────
def _btn(parent, text, font, bg, fg, hv_bg, hv_fg=None,
         pady=7, padx=14, state="normal", command=None):
    if hv_fg is None:
        hv_fg = fg
    b = tk.Button(parent, text=text, font=font,
                  bg=bg, fg=fg,
                  activebackground=hv_bg, activeforeground=hv_fg,
                  relief="flat", bd=0, cursor="hand2",
                  padx=padx, pady=pady,
                  state=state, command=command)
    b.bind("<Enter>", lambda e: b.config(bg=hv_bg, fg=hv_fg)
           if str(b["state"]) != "disabled" else None)
    b.bind("<Leave>", lambda e: b.config(bg=bg, fg=fg)
           if str(b["state"]) != "disabled" else None)
    return b


# ── Logo loader ────────────────────────────────────────────────────────────
def _load_logo(path, max_w, max_h, bg_hex=SURFACE):
    try:
        from PIL import Image, ImageTk
        img = Image.open(path).convert("RGBA")
        img.thumbnail((max_w, max_h), Image.LANCZOS)
        r, g, b = (int(bg_hex.lstrip("#")[i:i+2], 16) for i in (0, 2, 4))
        canvas = Image.new("RGBA", img.size, (r, g, b, 255))
        canvas.paste(img, mask=img.split()[3])
        return ImageTk.PhotoImage(canvas.convert("RGB"))
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════
class ComparadorApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title(f"{NOME_PROGRAMA} — SANEPAR / 2S Engenharia e Geotecnologia")
        try:
            self.iconbitmap(_res("assets", "oseaudit.ico"))
        except Exception:
            pass
        self.geometry("1020x860")
        self.minsize(860, 700)
        self.configure(bg=BG)
        self.resizable(True, True)

        self._var_mapas     = tk.StringVar()
        self._var_perfis    = tk.StringVar()
        self._var_planilhas = tk.StringVar()
        self._var_saida     = tk.StringVar()

        self._rodando     = False
        self._resultados  = []
        self._faltantes   = None

        # Tab state
        self._tab_frames     = {}
        self._tab_btns       = {}
        self._tab_inds       = {}
        self._sub_frames     = {}
        self._sub_btns       = {}
        self._sub_inds       = {}

        self._configure_styles()
        self._build_ui()
        self.withdraw()
        self._show_splash()

    # ──────────────────────────────────────────────────────────── Splash
    def _show_splash(self):
        SplashScreen(self, on_close=self._after_splash)

    def _after_splash(self):
        self.deiconify()
        self._centralizar()
        threading.Thread(target=self._check_updates, daemon=True).start()

    # ──────────────────────────────────────────────────────── Auto-update
    def _check_updates(self):
        try:
            url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
            req = urllib.request.Request(url, headers={"User-Agent": "OSEAudit"})
            with urllib.request.urlopen(req, timeout=8) as r:
                data = json.loads(r.read())
            tag = data.get("tag_name", "").lstrip("v")
            if not tag or tag == VERSAO:
                return
            asset = next(
                (a for a in data.get("assets", [])
                 if a["name"].lower().endswith(".exe")),
                None)
            if not asset:
                return
            self.after(0, lambda: self._start_update(tag, asset["browser_download_url"]))
        except Exception:
            pass

    def _start_update(self, nova_versao, download_url):
        """Inicia download silencioso em background — igual ao CodePro."""
        self._uc_version_lbl.config(text=f"v{nova_versao} disponível")
        self._uc_show("dl")
        threading.Thread(target=self._download_update_bg,
                         args=(download_url,), daemon=True).start()

    def _download_update_bg(self, url):
        try:
            tmp = tempfile.mktemp(suffix="_OSEAudit_Setup.exe")

            def _hook(count, block_size, total_size):
                if total_size > 0:
                    pct = min(100, int(count * block_size * 100 / total_size))
                    self.after(0, self._uc_set_progress, pct)

            urllib.request.urlretrieve(url, tmp, reporthook=_hook)
            self.after(0, self._uc_show, "ready")
            self.after(0, self._uc_start_countdown, tmp)
        except Exception:
            self.after(0, self._uc_hide)

    # ── Update card helpers ────────────────────────────────────────
    def _build_update_card(self):
        """Card flutuante no canto inferior direito (estilo CodePro)."""
        card = tk.Frame(self, bg=SURFACE,
                        highlightbackground="#1E3A5F",
                        highlightthickness=1)
        self._uc_card = card

        # ── Estado: baixando ───────────────────────────────────────
        dl = tk.Frame(card, bg=SURFACE, padx=14, pady=12)
        dl.pack(fill="x")
        self._uc_dl_frame = dl

        top = tk.Frame(dl, bg=SURFACE)
        top.pack(fill="x", pady=(0, 10))

        # Ícone
        icon_wrap = tk.Frame(top, bg="#0D1F3C",
                              highlightbackground="#1E3A5F", highlightthickness=1,
                              width=32, height=32)
        icon_wrap.pack(side="left", padx=(0, 10))
        icon_wrap.pack_propagate(False)
        self._uc_spin_lbl = tk.Label(icon_wrap, text="↻",
                                      font=("Segoe UI", 14), fg=BLU, bg="#0D1F3C")
        self._uc_spin_lbl.pack(expand=True)

        txt = tk.Frame(top, bg=SURFACE)
        txt.pack(side="left", fill="x", expand=True)
        tk.Label(txt, text="Baixando atualização",
                 font=F_BOLD, fg=TXT, bg=SURFACE).pack(anchor="w")
        self._uc_version_lbl = tk.Label(txt, text="",
                                         font=F_SMALL, fg=TXT2, bg=SURFACE)
        self._uc_version_lbl.pack(anchor="w")

        self._uc_pct_lbl = tk.Label(top, text="0%",
                                     font=("Consolas", 10, "bold"),
                                     fg=BLU, bg=SURFACE)
        self._uc_pct_lbl.pack(side="right")

        # Barra de progresso
        bar_bg = tk.Frame(dl, bg=CARD, height=4)
        bar_bg.pack(fill="x")
        bar_bg.pack_propagate(False)
        self._uc_bar_fill = tk.Frame(bar_bg, bg=BLU, height=4)
        self._uc_bar_fill.place(x=0, y=0, height=4, relwidth=0.0)

        # ── Estado: pronto para instalar ───────────────────────────
        rd = tk.Frame(card, bg=SURFACE, padx=14, pady=12)
        self._uc_rd_frame = rd

        top2 = tk.Frame(rd, bg=SURFACE)
        top2.pack(fill="x", pady=(0, 10))

        icon_wrap2 = tk.Frame(top2, bg="#0A2A1A",
                               highlightbackground="#14532D", highlightthickness=1,
                               width=32, height=32)
        icon_wrap2.pack(side="left", padx=(0, 10))
        icon_wrap2.pack_propagate(False)
        tk.Label(icon_wrap2, text="✓",
                 font=("Segoe UI", 13, "bold"), fg=GRN, bg="#0A2A1A").pack(expand=True)

        txt2 = tk.Frame(top2, bg=SURFACE)
        txt2.pack(side="left", fill="x", expand=True)
        tk.Label(txt2, text="Atualização baixada",
                 font=F_BOLD, fg=TXT, bg=SURFACE).pack(anchor="w")
        self._uc_countdown_lbl = tk.Label(txt2, text="Reiniciando em 5s…",
                                           font=F_SMALL, fg=TXT2, bg=SURFACE)
        self._uc_countdown_lbl.pack(anchor="w")

        self._uc_now_btn = _btn(top2, text="Agora",
                                font=("Segoe UI", 9, "bold"),
                                bg="#0A2A1A", fg=GRN, hv_bg="#14532D",
                                padx=8, pady=3)
        self._uc_now_btn.pack(side="right")

        # Barra de countdown (vai diminuindo)
        bar_bg2 = tk.Frame(rd, bg=CARD, height=4)
        bar_bg2.pack(fill="x")
        bar_bg2.pack_propagate(False)
        self._uc_cd_fill = tk.Frame(bar_bg2, bg=GRN, height=4)
        self._uc_cd_fill.place(x=0, y=0, height=4, relwidth=1.0)

    def _uc_show(self, state):
        """Exibe o card no canto inferior direito."""
        self._uc_dl_frame.pack_forget()
        self._uc_rd_frame.pack_forget()
        if state == "dl":
            self._uc_dl_frame.pack(fill="x")
        else:
            self._uc_rd_frame.pack(fill="x")
        # Posiciona acima da status bar (26px) com margem de 16px
        self._uc_card.place(relx=1.0, rely=1.0, anchor="se", x=-16, y=-42)
        self._uc_card.lift()

    def _uc_hide(self):
        self._uc_card.place_forget()

    def _uc_set_progress(self, pct):
        self._uc_pct_lbl.config(text=f"{pct}%")
        self._uc_bar_fill.place(x=0, y=0, height=4,
                                 relwidth=min(1.0, pct / 100))
        self._uc_spin_lbl.config(text="↻" if pct % 20 < 10 else "⟳")

    def _uc_start_countdown(self, tmp_path, secs=5):
        self._uc_now_btn.config(
            command=lambda: self._install_update(tmp_path))
        self._uc_do_countdown(tmp_path, secs)

    def _uc_do_countdown(self, tmp_path, secs):
        if secs <= 0:
            self._install_update(tmp_path)
            return
        self._uc_countdown_lbl.config(text=f"Reiniciando em {secs}s…")
        self._uc_cd_fill.place(x=0, y=0, height=4,
                                relwidth=max(0.0, secs / 5))
        self.after(1000, self._uc_do_countdown, tmp_path, secs - 1)

    def _install_update(self, tmp_path):
        try:
            subprocess.Popen([tmp_path], shell=True)
            self.destroy()
        except Exception:
            self._uc_hide()

    # ──────────────────────────────────────────────────────── ttk styles
    def _configure_styles(self):
        s = ttk.Style(self)
        s.theme_use("clam")

        # Treeview
        s.configure("Treeview",
                    background=CARD, foreground=TXT,
                    fieldbackground=CARD,
                    rowheight=28, font=F_MED, borderwidth=0)
        s.configure("Treeview.Heading",
                    background=SURFACE, foreground=TXT2,
                    font=("Segoe UI", 8, "bold"),
                    relief="flat", borderwidth=0)
        s.map("Treeview",
              background=[("selected", RED_D)],
              foreground=[("selected", RED_L)])
        s.map("Treeview.Heading",
              background=[("active", CARD)],
              relief=[("active", "flat")])

        # Progress bar — thin accent stripe
        s.configure("Red.Horizontal.TProgressbar",
                    troughcolor=CARD, background=RED,
                    thickness=3, borderwidth=0)

        # Scrollbars — minimal
        for orient in ("Vertical", "Horizontal"):
            s.configure(f"{orient}.TScrollbar",
                        background=OVERLAY, troughcolor=SURFACE,
                        arrowcolor=TXT3, bordercolor=SURFACE,
                        lightcolor=SURFACE, darkcolor=SURFACE,
                        width=8 if orient == "Vertical" else 8)
        s.map("TScrollbar", background=[("active", BORDER)])

    # ──────────────────────────────────────────────────────────── Build UI
    def _build_ui(self):
        self._build_header()
        self._build_folders_panel()
        self._build_tab_bar()
        self._content_area = tk.Frame(self, bg=BG)
        self._content_area.pack(fill="both", expand=True)
        self._build_tab_audit()
        self._build_tab_comp()
        self._build_statusbar()
        self._build_update_card()
        self._switch_tab("audit")

    # ════════════════════════════════════════════════ HEADER (dark unified)
    def _build_header(self):
        # Top accent bar
        tk.Frame(self, bg=RED, height=3).pack(fill="x")

        hdr = tk.Frame(self, bg=SURFACE)
        hdr.pack(fill="x")

        # Left — A2Z logo
        self._img_a2z = _load_logo(_res("assets", "logo_a2z.png"), 160, 60, SURFACE)
        left = tk.Frame(hdr, bg=SURFACE, width=196)
        left.pack(side="left", fill="y", padx=(24, 0), pady=14)
        left.pack_propagate(False)
        if self._img_a2z:
            tk.Label(left, image=self._img_a2z, bg=SURFACE, anchor="center").pack(expand=True)
        else:
            tk.Label(left, text="A2Z PROJETOS",
                     font=("Segoe UI", 13, "bold"), fg=RED, bg=SURFACE).pack(expand=True)

        # Right — 2S logo
        self._img_2s = _load_logo(_res("assets", "logo_2s.png"), 190, 60, SURFACE)
        right = tk.Frame(hdr, bg=SURFACE, width=216)
        right.pack(side="right", fill="y", padx=(0, 24), pady=14)
        right.pack_propagate(False)
        if self._img_2s:
            tk.Label(right, image=self._img_2s, bg=SURFACE, anchor="center").pack(expand=True)
        else:
            tk.Label(right, text="2S ENGENHARIA",
                     font=("Segoe UI", 11, "bold"), fg=TXT2, bg=SURFACE).pack(expand=True)

        # Center — title + subtitle + version pill
        ctr = tk.Frame(hdr, bg=SURFACE)
        ctr.pack(fill="both", expand=True, pady=12)

        tk.Label(ctr, text=NOME_PROGRAMA,
                 font=("Segoe UI", 22, "bold"),
                 fg=TXT, bg=SURFACE).pack()

        tk.Label(ctr, text=SUBTITULO,
                 font=("Segoe UI", 9), fg=TXT2, bg=SURFACE).pack(pady=(3, 0))

        pill_wrap = tk.Frame(ctr, bg=SURFACE)
        pill_wrap.pack(pady=(8, 0))
        pill = tk.Frame(pill_wrap, bg=RED_D, padx=12, pady=3)
        pill.pack()
        tk.Label(pill, text=f"v{VERSAO}",
                 font=("Segoe UI", 8, "bold"), fg=RED_L, bg=RED_D).pack()

        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

    # ════════════════════════════════════════════ FOLDER PANEL
    def _build_folders_panel(self):
        # Section cap label
        cap = tk.Frame(self, bg=BG, padx=20, pady=(10, 4))
        cap.pack(fill="x")
        tk.Label(cap, text="DOCUMENTOS",
                 font=F_CAP, fg=TXT3, bg=BG).pack(side="left")

        panel = tk.Frame(self, bg=SURFACE, padx=16, pady=10)
        panel.pack(fill="x", padx=20, pady=(0, 2))

        rows_cfg = [
            ("◈  MAPAS",     self._var_mapas,
             lambda: self._browse_folder(self._var_mapas,     "Selecione a pasta MAPAS")),
            ("◈  PERFIS",    self._var_perfis,
             lambda: self._browse_folder(self._var_perfis,    "Selecione a pasta PERFIS")),
            ("◈  PLANILHAS", self._var_planilhas,
             lambda: self._browse_folder(self._var_planilhas, "Selecione a pasta PLANILHAS")),
        ]

        for i, (label, var, cmd) in enumerate(rows_cfg):
            if i:
                tk.Frame(panel, bg=BORDER, height=1).pack(fill="x", pady=(2, 0))

            row = tk.Frame(panel, bg=SURFACE)
            row.pack(fill="x", pady=5)

            tk.Label(row, text=label, font=("Segoe UI", 9, "bold"),
                     fg=TXT2, bg=SURFACE, width=13, anchor="w").pack(side="left")

            ent = tk.Entry(row, textvariable=var, font=F_MED,
                           bg=CARD, fg=TXT, relief="flat", bd=0,
                           insertbackground=TXT,
                           selectbackground=RED_D, selectforeground=TXT)
            ent.pack(side="left", fill="x", expand=True, ipady=6, padx=(0, 8))

            b = _btn(row, text="···", font=("Segoe UI", 11, "bold"),
                     bg=OVERLAY, fg=TXT2, hv_bg=BORDER, hv_fg=TXT,
                     padx=10, pady=3, command=cmd)
            b.pack(side="left")

        tk.Frame(self, bg=BORDER, height=1).pack(fill="x", pady=(8, 0))

    def _browse_folder(self, var, title):
        p = filedialog.askdirectory(title=title)
        if p:
            var.set(p)

    # ═══════════════════════════════════════════════ CUSTOM TAB BAR
    def _build_tab_bar(self):
        bar = tk.Frame(self, bg=SURFACE)
        bar.pack(fill="x")

        for key, label in [("audit", "  ◈   Auditoria de Arquivos  "),
                            ("comp",  "  ◉   Comparação OSE  ")]:
            col = tk.Frame(bar, bg=SURFACE)
            col.pack(side="left")

            b = tk.Button(col, text=label, font=F_BOLD,
                          bg=SURFACE, fg=TXT3,
                          activebackground=SURFACE, activeforeground=TXT,
                          relief="flat", bd=0, cursor="hand2",
                          padx=2, pady=11,
                          command=lambda k=key: self._switch_tab(k))
            b.pack(fill="x")

            ind = tk.Frame(col, bg=SURFACE, height=2)
            ind.pack(fill="x")

            self._tab_btns[key] = b
            self._tab_inds[key] = ind

        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

    def _switch_tab(self, key):
        for k, fr in self._tab_frames.items():
            fr.pack_forget()
            self._tab_btns[k].config(fg=TXT3)
            self._tab_inds[k].config(bg=SURFACE)
        if key in self._tab_frames:
            self._tab_frames[key].pack(fill="both", expand=True)
        self._tab_btns[key].config(fg=TXT)
        self._tab_inds[key].config(bg=RED)

    # ════════════════════════════════════ TAB 1 — AUDITORIA DE ARQUIVOS
    def _build_tab_audit(self):
        frame = tk.Frame(self._content_area, bg=BG)
        self._tab_frames["audit"] = frame

        # Action bar
        ctrl = tk.Frame(frame, bg=BG, padx=20, pady=14)
        ctrl.pack(fill="x")

        self._btn_audit = _btn(ctrl,
            text="▶   VERIFICAR ARQUIVOS FALTANTES",
            font=("Segoe UI", 10, "bold"),
            bg=RED, fg="#FFFFFF", hv_bg=RED_H,
            pady=9, padx=20, command=self._executar_audit)
        self._btn_audit.pack(side="left")

        self._lbl_audit_status = tk.Label(ctrl, text="", font=F_UI, bg=BG, fg=TXT2)
        self._lbl_audit_status.pack(side="left", padx=16)

        self._btn_audit_export = _btn(ctrl,
            text="⬇  Exportar",
            font=F_BOLD, bg=OVERLAY, fg=TXT2, hv_bg=BORDER, hv_fg=TXT,
            pady=9, padx=14, state="disabled", command=self._export_audit)
        self._btn_audit_export.pack(side="right")

        tk.Frame(frame, bg=BORDER, height=1).pack(fill="x")

        # Summary bar
        sumbar = tk.Frame(frame, bg=SURFACE, padx=20, pady=10)
        sumbar.pack(fill="x")

        self._lbl_sum_mapa     = tk.Label(sumbar, text="MAPA: —",     font=F_MED, fg=TXT3, bg=SURFACE)
        self._lbl_sum_perfil   = tk.Label(sumbar, text="PERFIL: —",   font=F_MED, fg=TXT3, bg=SURFACE)
        self._lbl_sum_planilha = tk.Label(sumbar, text="PLANILHA: —", font=F_MED, fg=TXT3, bg=SURFACE)

        self._lbl_sum_mapa.pack(side="left", padx=(0, 32))
        self._lbl_sum_perfil.pack(side="left", padx=(0, 32))
        self._lbl_sum_planilha.pack(side="left")

        tk.Frame(frame, bg=BORDER, height=1).pack(fill="x")

        # Treeview
        tvf = tk.Frame(frame, bg=BG)
        tvf.pack(fill="both", expand=True)

        cols = ("arquivo", "mapa", "perfil", "planilha")
        self._tree_audit = ttk.Treeview(tvf, columns=cols,
                                        show="headings", selectmode="browse")

        self._tree_audit.heading("arquivo",  text="Arquivo",  anchor="w")
        self._tree_audit.heading("mapa",     text="MAPA",     anchor="center")
        self._tree_audit.heading("perfil",   text="PERFIL",   anchor="center")
        self._tree_audit.heading("planilha", text="PLANILHA", anchor="center")

        self._tree_audit.column("arquivo",  width=400, minwidth=200, anchor="w")
        self._tree_audit.column("mapa",     width=120, minwidth=80,  anchor="center")
        self._tree_audit.column("perfil",   width=120, minwidth=80,  anchor="center")
        self._tree_audit.column("planilha", width=120, minwidth=80,  anchor="center")

        vsb = ttk.Scrollbar(tvf, orient="vertical",   command=self._tree_audit.yview)
        hsb = ttk.Scrollbar(tvf, orient="horizontal", command=self._tree_audit.xview)
        self._tree_audit.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree_audit.pack(fill="both", expand=True)

        self._tree_audit.tag_configure("complete", background=GRN_BG,    foreground=GRN)
        self._tree_audit.tag_configure("missing",  background="#2D1515",  foreground=ERR)
        self._tree_audit.tag_configure("even",     background=CARD,       foreground=TXT2)
        self._tree_audit.tag_configure("odd",      background=SURFACE,    foreground=TXT2)

        self._tree_audit.insert("", "end",
            values=("—", "Clique em Verificar para iniciar", "—", "—"),
            tags=("even",))

    # ── Executar auditoria ─────────────────────────────────────────
    def _executar_audit(self):
        m  = self._var_mapas.get().strip()
        p  = self._var_perfis.get().strip()
        pl = self._var_planilhas.get().strip()

        faltam = [n for v, n in [(m, "MAPAS"), (p, "PERFIS"), (pl, "PLANILHAS")] if not v]
        if faltam:
            messagebox.showwarning("Atenção",
                "Informe as pastas:\n• " + "\n• ".join(faltam)); return
        for path, name in [(m, "MAPAS"), (p, "PERFIS"), (pl, "PLANILHAS")]:
            if not os.path.isdir(path):
                messagebox.showerror("Erro", f"Pasta {name} não encontrada:\n{path}"); return

        self._btn_audit.config(state="disabled", text="⏳  Verificando…")
        self._lbl_audit_status.config(text="Verificando…", fg=TXT2)
        self._btn_audit_export.config(state="disabled")
        for lbl, t in [(self._lbl_sum_mapa,     "MAPA: …"),
                       (self._lbl_sum_perfil,   "PERFIL: …"),
                       (self._lbl_sum_planilha, "PLANILHA: …")]:
            lbl.config(text=t, fg=TXT3)
        self._set_status("Verificando arquivos…", "info")
        threading.Thread(target=self._run_audit_bg, args=(m, p, pl), daemon=True).start()

    def _run_audit_bg(self, pm, pp, ppl):
        try:
            result = verificar_faltantes(pm, pp, ppl)
            self.after(0, self._audit_ok, result)
        except Exception as exc:
            self.after(0, self._audit_err, str(exc))

    def _audit_ok(self, result):
        self._faltantes = result
        self._btn_audit.config(state="normal", text="▶   VERIFICAR ARQUIVOS FALTANTES")

        fm  = len(result["falta_mapa"])
        fp  = len(result["falta_perfil"])
        fpl = len(result["falta_planilha"])
        tot = len(result["todos"])

        def _slbl(lbl, name, count):
            if count == 0:
                lbl.config(text=f"✅  {name}: Completa", fg=GRN)
            else:
                lbl.config(text=f"⚠   {name}: {count} faltando", fg=AMB)

        _slbl(self._lbl_sum_mapa,     "MAPA",     fm)
        _slbl(self._lbl_sum_perfil,   "PERFIL",   fp)
        _slbl(self._lbl_sum_planilha, "PLANILHA", fpl)

        total_miss = fm + fp + fpl
        if total_miss == 0:
            self._lbl_audit_status.config(text=f"✅  {tot} arquivo(s) — completo.", fg=GRN)
            self._set_status(f"Auditoria concluída — {tot} arquivo(s), nenhuma falta", "ok")
        else:
            self._lbl_audit_status.config(
                text=f"⚠   {tot} arquivo(s) — {total_miss} ausência(s).", fg=AMB)
            self._set_status(
                f"Auditoria concluída — {total_miss} arquivo(s) faltando", "warn")

        for item in self._tree_audit.get_children():
            self._tree_audit.delete(item)

        set_m  = set(result["mapa"])
        set_p  = set(result["perfil"])
        set_pl = set(result["planilha"])

        for nome in result["todos"]:
            hm  = "✅" if nome in set_m  else "❌"
            hp  = "✅" if nome in set_p  else "❌"
            hpl = "✅" if nome in set_pl else "❌"
            tag = "missing" if "❌" in (hm, hp, hpl) else "complete"
            self._tree_audit.insert("", "end",
                values=(nome, hm, hp, hpl), tags=(tag,))

        self._btn_audit_export.config(state="normal")

    def _audit_err(self, erro):
        self._btn_audit.config(state="normal", text="▶   VERIFICAR ARQUIVOS FALTANTES")
        self._lbl_audit_status.config(text=f"❌  {erro}", fg=ERR)
        self._set_status(f"Erro na auditoria: {erro}", "err")
        messagebox.showerror("Erro", erro)

    # ── Exportar auditoria ─────────────────────────────────────────
    def _export_audit(self):
        if not self._faltantes:
            return
        fmt = self._ask_format(["Excel (.xlsx)", "TXT (.txt)"])
        if not fmt:
            return
        ext   = ".xlsx" if "Excel" in fmt else ".txt"
        label = "Excel" if "Excel" in fmt else "TXT"
        path  = filedialog.asksaveasfilename(
            title=f"Salvar auditoria {label}",
            defaultextension=ext,
            filetypes=[(label, f"*{ext}"), ("Todos", "*.*")])
        if not path:
            return
        try:
            if label == "Excel":
                self._export_audit_excel(path)
            else:
                self._export_audit_txt(path)
            if messagebox.askyesno("Exportado", "Relatório exportado!\n\nDeseja abrir?"):
                os.startfile(path)
        except Exception as exc:
            messagebox.showerror("Erro", str(exc))

    def _export_audit_txt(self, path):
        r  = self._faltantes
        ts = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        lines = ["=" * 62,
                 f"  {NOME_PROGRAMA} — Auditoria de Arquivos Faltantes",
                 f"  Gerado em: {ts}", "=" * 62,
                 f"\nTotal de arquivos identificados: {len(r['todos'])}"]
        for chave, nome in [("falta_mapa", "MAPA"),
                             ("falta_perfil", "PERFIL"),
                             ("falta_planilha", "PLANILHA")]:
            lines.append(f"\n{nome}:")
            if r[chave]:
                for f in r[chave]:
                    lines.append(f"  ❌  {f}")
            else:
                lines.append("  ✓  Completa")
            lines.append(f"  Total faltando: {len(r[chave])}")
        lines += ["\n" + "=" * 62,
                  f"  {NOME_PROGRAMA} — A2Z Projetos + 2S Engenharia e Geotecnologia",
                  "=" * 62]
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    def _export_audit_excel(self, path):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Auditoria de Arquivos"

        def _bd():
            s = Side(style="thin", color="444444")
            return Border(left=s, right=s, top=s, bottom=s)

        def _hdr(row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill      = PatternFill("solid", start_color="8B0000")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _bd()

        def _cell(row, col, val, bg="1C1C1C", fg="FFFFFF"):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Arial", color=fg, size=9)
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(
                horizontal="left" if col == 1 else "center",
                vertical="center")
            c.border = _bd()

        ws.merge_cells("A1:D1")
        t = ws["A1"]
        t.value     = f"{NOME_PROGRAMA} — Auditoria de Arquivos"
        t.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        t.fill      = PatternFill("solid", start_color="1A0000")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        for col, lbl in enumerate(["Arquivo", "MAPA", "PERFIL", "PLANILHA"], 1):
            _hdr(2, col, lbl)
        ws.row_dimensions[2].height = 20

        r   = self._faltantes
        sm  = set(r["mapa"])
        sp  = set(r["perfil"])
        spl = set(r["planilha"])

        for i, nome in enumerate(r["todos"]):
            row = i + 3
            hm  = "✅" if nome in sm  else "❌"
            hp  = "✅" if nome in sp  else "❌"
            hpl = "✅" if nome in spl else "❌"
            miss = "❌" in (hm, hp, hpl)
            bg = "1E0505" if miss else "071507"
            fg = "FF6B6B" if miss else "4CD964"
            _cell(row, 1, nome, bg, fg)
            for col, val in zip([2, 3, 4], [hm, hp, hpl]):
                _cell(row, col, val, bg, fg)
            ws.row_dimensions[row].height = 16

        ws.column_dimensions["A"].width = 38
        for col in ["B", "C", "D"]:
            ws.column_dimensions[col].width = 14
        ws.freeze_panes = "A3"
        wb.save(path)

    # ════════════════════════════════════════ TAB 2 — COMPARAÇÃO OSE
    def _build_tab_comp(self):
        frame = tk.Frame(self._content_area, bg=BG)
        self._tab_frames["comp"] = frame

        # Output file row
        out_row = tk.Frame(frame, bg=SURFACE, padx=20, pady=10)
        out_row.pack(fill="x")

        tk.Label(out_row, text="💾  Relatório de saída:",
                 font=F_BOLD, fg=TXT2, bg=SURFACE).pack(side="left")

        ent = tk.Entry(out_row, textvariable=self._var_saida, font=F_UI,
                       bg=CARD, fg=TXT, relief="flat", bd=0,
                       insertbackground=TXT, selectbackground=RED_D)
        ent.pack(side="left", fill="x", expand=True, ipady=6, padx=10)

        bs = _btn(out_row, text="···", font=("Segoe UI", 11, "bold"),
                  bg=OVERLAY, fg=TXT2, hv_bg=BORDER, hv_fg=TXT,
                  padx=10, pady=3, command=self._browse_saida)
        bs.pack(side="left")

        tk.Frame(frame, bg=BORDER, height=1).pack(fill="x")

        # Run + progress
        ctrl = tk.Frame(frame, bg=BG, padx=20, pady=14)
        ctrl.pack(fill="x")

        self._btn_run = _btn(ctrl,
            text="▶   EXECUTAR COMPARAÇÃO OSE",
            font=("Segoe UI", 11, "bold"),
            bg=RED, fg="#FFFFFF", hv_bg=RED_H,
            pady=10, padx=24, command=self._executar_comp)
        self._btn_run.pack(side="left")

        prog_wrap = tk.Frame(ctrl, bg=BG)
        prog_wrap.pack(side="left", padx=16, fill="x", expand=True)

        self._progress = ttk.Progressbar(prog_wrap,
            style="Red.Horizontal.TProgressbar",
            mode="indeterminate", length=300)
        self._progress.pack(fill="x")

        self._lbl_status = tk.Label(prog_wrap, text="", font=F_SMALL, bg=BG, fg=TXT2)
        self._lbl_status.pack(anchor="w", pady=(4, 0))

        self._btn_export = _btn(ctrl,
            text="⬇  Exportar",
            font=F_BOLD, bg=OVERLAY, fg=TXT2, hv_bg=BORDER, hv_fg=TXT,
            pady=10, padx=14, state="disabled", command=self._export_report)
        self._btn_export.pack(side="right")

        tk.Frame(frame, bg=BORDER, height=1).pack(fill="x")

        # Sub-tabs: Log + Relatório
        sub_bar = tk.Frame(frame, bg=SURFACE)
        sub_bar.pack(fill="x")

        for key, label in [("log", "  Log de Execução  "),
                            ("rep", "  Relatório de Erros  ")]:
            col = tk.Frame(sub_bar, bg=SURFACE)
            col.pack(side="left")

            b = tk.Button(col, text=label, font=("Segoe UI", 9, "bold"),
                          bg=SURFACE, fg=TXT3,
                          activebackground=SURFACE, activeforeground=TXT,
                          relief="flat", bd=0, cursor="hand2",
                          padx=2, pady=8,
                          command=lambda k=key: self._switch_sub(k))
            b.pack(fill="x")

            ind = tk.Frame(col, bg=SURFACE, height=2)
            ind.pack(fill="x")

            self._sub_btns[key] = b
            self._sub_inds[key] = ind

        tk.Frame(frame, bg=BORDER, height=1).pack(fill="x")

        sub_content = tk.Frame(frame, bg=BG)
        sub_content.pack(fill="both", expand=True)
        self._sub_content = sub_content

        self._tab_log = tk.Frame(sub_content, bg=LOG_BG)
        self._tab_rep = tk.Frame(sub_content, bg=BG)
        self._sub_frames["log"] = self._tab_log
        self._sub_frames["rep"] = self._tab_rep

        self._build_log_tab()
        self._build_report_tab()
        self._switch_sub("log")

    def _switch_sub(self, key):
        for k, fr in self._sub_frames.items():
            fr.pack_forget()
            self._sub_btns[k].config(fg=TXT3)
            self._sub_inds[k].config(bg=SURFACE)
        self._sub_frames[key].pack(fill="both", expand=True)
        self._sub_btns[key].config(fg=TXT)
        self._sub_inds[key].config(bg=BORDER_A)

    # ── Browse saída ───────────────────────────────────────────────
    def _browse_saida(self):
        ini = self._var_mapas.get() or "/"
        arq = filedialog.asksaveasfilename(
            title="Salvar relatório como…",
            initialdir=ini, defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])
        if arq:
            self._var_saida.set(arq)

    # ── Executar comparação ────────────────────────────────────────
    def _executar_comp(self):
        if self._rodando:
            return
        m  = self._var_mapas.get().strip()
        p  = self._var_perfis.get().strip()
        pl = self._var_planilhas.get().strip()

        faltam = [n for v, n in [(m, "MAPAS"), (p, "PERFIS"), (pl, "PLANILHAS")] if not v]
        if faltam:
            messagebox.showwarning("Atenção",
                "Informe as pastas:\n• " + "\n• ".join(faltam)); return
        for path, name in [(m, "MAPAS"), (p, "PERFIS"), (pl, "PLANILHAS")]:
            if not os.path.isdir(path):
                messagebox.showerror("Erro", f"Pasta {name} não encontrada:\n{path}"); return

        saida = self._var_saida.get().strip()
        if not saida:
            ts    = datetime.now().strftime("%Y%m%d_%H%M")
            saida = os.path.join(m, f"relatorio_divergencias_{ts}.xlsx")
            self._var_saida.set(saida)

        self._rodando    = True
        self._resultados = []
        self._btn_run.config(state="disabled", text="⏳  Processando…")
        self._progress.start(12)
        self._lbl_status.config(text="Processando…", fg=TXT2)
        self._btn_export.config(state="disabled")
        self._log_clear()
        self._set_status("Comparação em andamento…", "info")

        self._log_write("Iniciando comparação\n", "info")
        for label, val in [("MAPAS", m), ("PERFIS", p),
                            ("PLANILHAS", pl), ("Saída", saida)]:
            self._log_write(f"{label:<12}: {val}\n", "dim")
        self._log_write("\n", "dim")

        threading.Thread(target=self._run_comp_bg,
                         args=(m, p, pl, saida), daemon=True).start()

    def _run_comp_bg(self, pm, pp, ppl, saida):
        buf = io.StringIO()
        try:
            with contextlib.redirect_stdout(buf):
                resultados = processar_pastas(pm, pp, ppl, saida)
            self.after(0, self._comp_ok, buf.getvalue(), saida, resultados or [])
        except Exception as exc:
            import traceback
            buf.write("\n" + traceback.format_exc())
            self.after(0, self._comp_err, buf.getvalue(), str(exc))

    def _comp_ok(self, log_txt, saida, resultados):
        self._rodando    = False
        self._resultados = resultados
        self._progress.stop()
        self._btn_run.config(state="normal", text="▶   EXECUTAR COMPARAÇÃO OSE")
        self._lbl_status.config(text="Concluído com sucesso.", fg=GRN)
        self._log_from_text(log_txt)
        self._log_write(f"\nRelatório salvo em:\n{saida}\n", "ok")
        self._update_report_tab(resultados)
        self._switch_sub("rep")
        total = sum(len(d) for *_, d in resultados)
        self._set_status(
            f"Concluído — {total} divergência(s) em {len(resultados)} OSE(s)",
            "ok" if total == 0 else "warn")

        if messagebox.askyesno("Concluído",
                               f"Comparação finalizada!\n\nAbrir relatório Excel?\n{saida}"):
            try:
                os.startfile(saida)
            except Exception:
                pass

    def _comp_err(self, log_txt, erro):
        self._rodando = False
        self._progress.stop()
        self._btn_run.config(state="normal", text="▶   EXECUTAR COMPARAÇÃO OSE")
        self._lbl_status.config(text="Erro no processamento.", fg=ERR)
        self._log_from_text(log_txt)
        self._log_write(f"\nERRO: {erro}\n", "err")
        self._set_status(f"Erro: {erro}", "err")
        messagebox.showerror("Erro", f"Ocorreu um erro:\n\n{erro}")

    # ── Log tab ────────────────────────────────────────────────────
    def _build_log_tab(self):
        self._log = tk.Text(
            self._tab_log, font=F_MONO, bg=LOG_BG, fg=LOG_FG,
            state="disabled", wrap="word", bd=0, padx=14, pady=10,
            selectbackground=CARD, insertbackground=TXT)
        sb = ttk.Scrollbar(self._tab_log, command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._log.pack(fill="both", expand=True)

        for tag, fg in [("ok",   GRN), ("err",  ERR),
                        ("warn", AMB), ("info", BLU),
                        ("dim",  TXT3)]:
            self._log.tag_config(tag, foreground=fg)

        self._log_write("Sistema pronto. Informe as pastas e clique em Executar.\n", "dim")

    # ── Report tab ─────────────────────────────────────────────────
    def _build_report_tab(self):
        top = tk.Frame(self._tab_rep, bg=SURFACE, padx=16, pady=10)
        top.pack(fill="x")

        self._lbl_summary = tk.Label(top,
            text="Nenhuma comparação executada ainda.",
            font=F_UI, fg=TXT2, bg=SURFACE, anchor="w")
        self._lbl_summary.pack(side="left", fill="x", expand=True)

        tk.Frame(self._tab_rep, bg=BORDER, height=1).pack(fill="x")

        tvf = tk.Frame(self._tab_rep, bg=BG)
        tvf.pack(fill="both", expand=True)

        cols = ("ose", "campo", "fonte_a", "valor_a",
                "fonte_b", "valor_b", "diferenca")
        self._tree = ttk.Treeview(tvf, columns=cols,
                                  show="headings", selectmode="browse")

        for cid, lbl, w, anch in [
            ("ose",       "OSE",       82,  "center"),
            ("campo",     "Campo",     185, "w"),
            ("fonte_a",   "Fonte A",   80,  "center"),
            ("valor_a",   "Valor A",   118, "center"),
            ("fonte_b",   "Fonte B",   80,  "center"),
            ("valor_b",   "Valor B",   118, "center"),
            ("diferenca", "Diferença", 88,  "center"),
        ]:
            self._tree.heading(cid, text=lbl, anchor=anch)
            self._tree.column(cid, width=w, minwidth=50, anchor=anch)

        vsb = ttk.Scrollbar(tvf, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(tvf, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree.pack(fill="both", expand=True)

        self._tree.tag_configure("even",   background=CARD,    foreground=TXT)
        self._tree.tag_configure("odd",    background=SURFACE, foreground=TXT)
        self._tree.tag_configure("ok_row", background=GRN_BG,  foreground=GRN)

        self._tree.insert("", "end",
            values=("—", "Execute uma comparação para ver os resultados",
                    "—", "—", "—", "—", "—"),
            tags=("even",))

    # ── Atualizar relatório ────────────────────────────────────────
    def _update_report_tab(self, resultados):
        for item in self._tree.get_children():
            self._tree.delete(item)

        total_div    = 0
        oses_com_div = 0
        idx          = 0

        for num, planilha, mapa, perfil, divs in resultados:
            if divs:
                oses_com_div += 1
                for div in divs:
                    tag = "even" if idx % 2 == 0 else "odd"
                    self._tree.insert("", "end", values=(
                        f"OSE-{num}", div.campo,
                        div.fonte_a, div.valor_a,
                        div.fonte_b, div.valor_b,
                        div.diferenca or "—",
                    ), tags=(tag,))
                    total_div += 1; idx += 1
            else:
                self._tree.insert("", "end", values=(
                    f"OSE-{num}", "✅  Sem divergências",
                    "—", "—", "—", "—", "—",
                ), tags=("ok_row",))
                idx += 1

        total_oses = len(resultados)
        if total_div == 0:
            self._lbl_summary.config(
                text=f"✅  Nenhuma divergência em {total_oses} OSE(s).", fg=GRN)
        else:
            self._lbl_summary.config(
                text=(f"⚠   {total_div} divergência(s) em "
                      f"{oses_com_div} de {total_oses} OSE(s)."),
                fg=AMB)
        self._btn_export.config(state="normal")

    # ── Exportar relatório OSE ─────────────────────────────────────
    def _export_report(self):
        if not self._resultados:
            return
        fmt = self._ask_format(["Excel (.xlsx)", "TXT (.txt)", "PDF (.pdf)"])
        if not fmt:
            return
        label = fmt.split()[0]
        ext   = {"Excel": ".xlsx", "TXT": ".txt", "PDF": ".pdf"}[label]
        path  = filedialog.asksaveasfilename(
            title=f"Salvar relatório {label}",
            defaultextension=ext,
            filetypes=[(label, f"*{ext}"), ("Todos", "*.*")])
        if not path:
            return
        try:
            {"Excel": self._do_export_excel,
             "TXT":   self._do_export_txt,
             "PDF":   self._do_export_pdf}[label](path)
            if messagebox.askyesno("Exportado",
                                   "Relatório exportado!\n\nDeseja abrir?"):
                os.startfile(path)
        except Exception as exc:
            messagebox.showerror("Erro ao Exportar", str(exc))

    # ── Format picker ──────────────────────────────────────────────
    def _ask_format(self, options):
        dlg = tk.Toplevel(self)
        dlg.title("Formato de Exportação")
        dlg.configure(bg=SURFACE)
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()
        h = 130 + len(options) * 32
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        dlg.geometry(f"300x{h}+{(sw-300)//2}+{(sh-h)//2}")

        tk.Frame(dlg, bg=RED, height=2).pack(fill="x")

        tk.Label(dlg, text="Escolha o formato:",
                 font=F_BOLD, fg=TXT, bg=SURFACE).pack(pady=(16, 10))

        chosen = tk.StringVar(value=options[0])
        for opt in options:
            tk.Radiobutton(dlg, text=opt, variable=chosen, value=opt,
                font=F_UI, fg=TXT, bg=SURFACE,
                activebackground=SURFACE, activeforeground=TXT,
                selectcolor=CARD, cursor="hand2").pack(anchor="w", padx=30)

        result = [None]

        def confirm():
            result[0] = chosen.get()
            dlg.destroy()

        bf = tk.Frame(dlg, bg=SURFACE)
        bf.pack(pady=14)
        _btn(bf, text="Exportar", font=F_BOLD,
             bg=RED, fg="#FFFFFF", hv_bg=RED_H,
             padx=14, pady=5, command=confirm).pack(side="left", padx=4)
        _btn(bf, text="Cancelar", font=F_UI,
             bg=OVERLAY, fg=TXT2, hv_bg=BORDER, hv_fg=TXT,
             padx=14, pady=5, command=dlg.destroy).pack(side="left", padx=4)

        dlg.wait_window()
        return result[0]

    # ── Excel export ───────────────────────────────────────────────
    def _do_export_excel(self, path):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for num, plan, mapa, perf, divs in self._resultados:
            gerar_aba_ose(wb, num, plan, mapa, perf, divs)
        gerar_aba_resumo(wb, self._resultados)
        wb.save(path)

    # ── TXT export ─────────────────────────────────────────────────
    def _do_export_txt(self, path):
        ts = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        lines = ["=" * 64,
                 f"  {NOME_PROGRAMA} — Relatório de Divergências",
                 f"  Gerado em: {ts}", "=" * 64]
        for num, plan, mapa, perf, divs in self._resultados:
            lines += [f"\nOSE-{num}", "-" * 42]
            if not divs:
                lines.append("  ✓ Nenhuma divergência")
            else:
                for d in divs:
                    lines += [f"  Campo : {d.campo}",
                               f"    {d.fonte_a:<12}: {d.valor_a}",
                               f"    {d.fonte_b:<12}: {d.valor_b}"]
                    if d.diferenca:
                        lines.append(f"    Diferença   : {d.diferenca}")
                    lines.append("")
        total = sum(len(d) for *_, d in self._resultados)
        lines += ["=" * 64,
                  f"  Total: {total} divergência(s) em {len(self._resultados)} OSE(s)",
                  f"  {NOME_PROGRAMA} — A2Z Projetos + 2S Engenharia", "=" * 64]
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

    # ── PDF export ─────────────────────────────────────────────────
    def _do_export_pdf(self, path):
        try:
            from fpdf import FPDF
        except ImportError:
            raise RuntimeError("fpdf2 não encontrada. Execute: pip install fpdf2")

        ts  = datetime.now().strftime("%d/%m/%Y %H:%M")
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        la = _res("assets", "logo_a2z.png")
        l2 = _res("assets", "logo_2s.png")
        y0 = 10
        if os.path.exists(la):
            try: pdf.image(la, x=10, y=y0, h=14)
            except Exception: pass
        if os.path.exists(l2):
            try: pdf.image(l2, x=163, y=y0+1, h=11)
            except Exception: pass

        pdf.set_y(y0)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(192, 0, 0)
        pdf.cell(0, 7, NOME_PROGRAMA, ln=False, align="C")
        pdf.ln(7)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(110, 110, 110)
        pdf.cell(0, 5, f"Relatorio de Divergencias — {ts}", ln=True, align="C")
        pdf.cell(0, 4, "A2Z Projetos em parceria com 2S Engenharia e Geotecnologia",
                  ln=True, align="C")
        pdf.ln(3)
        pdf.set_draw_color(192, 0, 0)
        pdf.set_line_width(0.5)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(4)

        cw = [20, 52, 22, 30, 22, 30, 24]
        for num, plan, mapa, perf, divs in self._resultados:
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(192, 0, 0)
            pdf.cell(0, 7, f"OSE-{num}", ln=True)
            if not divs:
                pdf.set_font("Helvetica", "", 9)
                pdf.set_text_color(0, 140, 0)
                pdf.cell(0, 6, "  Nenhuma divergencia", ln=True)
            else:
                pdf.set_font("Helvetica", "B", 7)
                pdf.set_fill_color(139, 0, 0)
                pdf.set_text_color(255, 255, 255)
                for lbl, w in zip(["OSE", "Campo", "Fte A", "Val A",
                                    "Fte B", "Val B", "Dif."], cw):
                    pdf.cell(w, 6, lbl, border=1, fill=True, align="C")
                pdf.ln()
                pdf.set_font("Helvetica", "", 7)
                for i, d in enumerate(divs):
                    fl = i % 2 == 0
                    pdf.set_fill_color(255, 240, 240) if fl \
                        else pdf.set_fill_color(255, 255, 255)
                    pdf.set_text_color(40, 40, 40)
                    for val, w in zip([f"OSE-{num}", d.campo, d.fonte_a,
                                       d.valor_a, d.fonte_b, d.valor_b,
                                       d.diferenca or "-"], cw):
                        pdf.cell(w, 5, str(val)[:18], border=1, fill=fl, align="C")
                    pdf.ln()
            pdf.ln(3)

        pdf.set_draw_color(192, 0, 0)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(40, 40, 40)
        total = sum(len(d) for *_, d in self._resultados)
        pdf.cell(0, 5,
                 f"Total: {total} divergencia(s) em {len(self._resultados)} OSE(s)",
                 ln=True)
        pdf.output(path)

    # ── Status bar ─────────────────────────────────────────────────
    def _build_statusbar(self):
        bar = tk.Frame(self, bg=SURFACE, height=26)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)

        tk.Frame(bar, bg=BORDER, height=1).pack(fill="x", side="top")

        self._sb_dot  = tk.Label(bar, text="●", font=F_SMALL, bg=SURFACE, fg=GRN)
        self._sb_text = tk.Label(bar, text="Pronto",
                                 font=F_SMALL, bg=SURFACE, fg=TXT2)
        sep   = tk.Label(bar, text=" │ ", font=F_SMALL, bg=SURFACE, fg=TXT3)
        cred  = tk.Label(bar,
                         text="A2Z Projetos × 2S Engenharia e Geotecnologia",
                         font=F_SMALL, bg=SURFACE, fg=TXT3)
        ver   = tk.Label(bar, text=f"v{VERSAO}", font=F_SMALL, bg=SURFACE, fg=TXT3)

        self._sb_dot.pack(side="left",  padx=(12, 4))
        self._sb_text.pack(side="left")
        sep.pack(side="left",  padx=4)
        cred.pack(side="left")
        ver.pack(side="right", padx=12)

    def _set_status(self, msg, kind="dim"):
        color = {"ok": GRN, "warn": AMB, "err": ERR,
                 "info": BLU, "dim": TXT3}.get(kind, TXT2)
        if hasattr(self, "_sb_dot"):
            self._sb_dot.config(fg=color)
            self._sb_text.config(text=msg, fg=color)

    # ── Log helpers ────────────────────────────────────────────────
    def _log_write(self, txt, tag=""):
        self._log.config(state="normal")
        self._log.insert("end", txt, tag)
        self._log.see("end")
        self._log.config(state="disabled")

    def _log_clear(self):
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")

    def _log_from_text(self, txt):
        for linha in txt.splitlines():
            if "✅" in linha or "Relatório salvo" in linha:
                tag = "ok"
            elif "⚠️" in linha or "divergência" in linha.lower():
                tag = "warn"
            elif "[ERRO]" in linha or "Erro" in linha:
                tag = "err"
            elif "OSEs" in linha or "Processando" in linha:
                tag = "info"
            else:
                tag = "dim"
            self._log_write(linha + "\n", tag)

    # ── Utils ──────────────────────────────────────────────────────
    def _centralizar(self):
        self.update_idletasks()
        w, h   = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw - w)//2}+{(sh - h)//2}")


if __name__ == "__main__":
    app = ComparadorApp()
    app.mainloop()
